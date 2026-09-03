from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from services.excel_orders import find_header_row, normalize_header, select_order_sheets


STATUS_HEADERS = {"STATUS", "STATUS ORDER", "ORDER STATUS"}
NEW_SN_HEADERS = {"SN ONT NEW", "SN ONT BARU", "SN NEW", "SN BARU"}
TICKET_HEADERS = {"TIKET", "TIKET ID", "TICKET", "TICKET ID", "INC", "NO TIKET"}
SERVICE_HEADERS = {"INET", "NO INET", "NO SERVICE", "NO INTERNET", "SERVICE NUMBER"}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def update_order_excel(
    file_path: Path,
    *,
    ticket_id: str = "",
    service_number: str = "",
    new_sn: str,
    status: str = "CLOSE",
) -> int:
    """Perbarui SN ONT NEW dan STATUS pada baris order yang cocok."""
    if not file_path.exists():
        raise FileNotFoundError(f"File Excel tidak ditemukan: {file_path.name}")

    workbook = load_workbook(file_path, read_only=False, data_only=False)
    updated = 0

    try:
        for sheet in select_order_sheets(workbook):
            try:
                header_row, _ = find_header_row(sheet)
            except ValueError:
                continue

            columns: dict[str, int] = {}
            for index, cell in enumerate(sheet[header_row], start=1):
                header = normalize_header(cell.value)
                if header in STATUS_HEADERS:
                    columns["status"] = index
                elif header in NEW_SN_HEADERS:
                    columns["new_sn"] = index
                elif header in TICKET_HEADERS:
                    columns["ticket_id"] = index
                elif header in SERVICE_HEADERS:
                    columns["service_number"] = index

            if "new_sn" not in columns or "status" not in columns:
                continue

            for row_index in range(header_row + 1, (sheet.max_row or header_row) + 1):
                row_ticket = _text(sheet.cell(row_index, columns.get("ticket_id", 1)).value) if "ticket_id" in columns else ""
                row_service = _text(sheet.cell(row_index, columns.get("service_number", 1)).value) if "service_number" in columns else ""

                ticket_match = bool(ticket_id and row_ticket.upper() == ticket_id.strip().upper())
                service_match = bool(service_number and row_service == service_number.strip())
                if not ticket_match and not service_match:
                    continue

                sheet.cell(row_index, columns["new_sn"]).value = new_sn.strip().upper()
                sheet.cell(row_index, columns["status"]).value = status.strip().upper()
                updated += 1

        if updated == 0:
            raise ValueError(
                "Baris order tidak ditemukan atau kolom STATUS/SN ONT NEW tidak tersedia."
            )

        workbook.save(file_path)
        return updated
    finally:
        workbook.close()
